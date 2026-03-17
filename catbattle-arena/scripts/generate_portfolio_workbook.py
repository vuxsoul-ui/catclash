from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("/Users/charon/go/catbattle-arena/output/spreadsheet/portfolio-management-2026-03-15.xlsx")

AS_OF_LABEL = "As of Friday, March 13, 2026 U.S. market close"
ACCOUNT_NAME = "PFFK"
PORTFOLIO_GOAL = "Maximize upside while reducing concentration risk and preserving covered-call flexibility."

HOLDINGS = [
    {
        "ticker": "VOO",
        "name": "Vanguard S&P 500 ETF",
        "shares": 4.99,
        "price": 552.75,
        "display_pl": -83.56,
        "display_pl_note": "Screenshot app P/L display",
        "target_phase1": 0.30,
        "target_phase2": 0.32,
        "covered_min_shares": 0,
        "action_now": "Accumulate on dips with trim proceeds; core holding.",
    },
    {
        "ticker": "ONDS",
        "name": "Ondas Holdings",
        "shares": 453.89,
        "price": 10.16,
        "display_pl": -403.19,
        "display_pl_note": "Screenshot app P/L display",
        "target_phase1": 0.22,
        "target_phase2": 0.18,
        "covered_min_shares": 300,
        "action_now": "Do not go below 300 shares until covered calls are closed or rolled.",
    },
    {
        "ticker": "NVDA",
        "name": "NVIDIA",
        "shares": 54.26,
        "price": 121.67,
        "display_pl": 46.26,
        "display_pl_note": "Screenshot app P/L display",
        "target_phase1": 0.38,
        "target_phase2": 0.40,
        "covered_min_shares": 0,
        "action_now": "Best trim source because it is the largest position and still a strong core name after trimming.",
    },
    {
        "ticker": "SLS",
        "name": "SELLAS Life Sciences",
        "shares": 57.78,
        "price": 1.75,
        "display_pl": 45.74,
        "display_pl_note": "Screenshot app P/L display",
        "target_phase1": 0.00,
        "target_phase2": 0.00,
        "covered_min_shares": 0,
        "action_now": "Treat as optionality only; exit on strength or simplify now.",
    },
]

OPTIONS = [
    {
        "underlying": "ONDS",
        "description": "ONDS $20 Call",
        "expiration": "2026-06-18",
        "contracts": -1,
        "strike": 20.00,
        "bid": 0.50,
        "ask": 0.54,
        "mid": 0.52,
        "last": 0.52,
        "delta": 0.2004,
        "iv": 1.1475,
        "open_interest": 17190,
        "volume": 828,
        "display_pl": 141.00,
        "buyback_trigger": "Buy back if mark <= $0.12 and you want to reopen upside with a higher strike.",
        "roll_trigger": "Roll up/out if ONDS closes above $14 or call delta rises above 0.35.",
        "action_now": "Hold. Strike is 96.9% above spot and premium left is still meaningful.",
    },
    {
        "underlying": "ONDS",
        "description": "ONDS $20 Call",
        "expiration": "2026-09-18",
        "contracts": -2,
        "strike": 20.00,
        "bid": 1.13,
        "ask": 1.27,
        "mid": 1.20,
        "last": 1.18,
        "delta": 0.3264,
        "iv": 1.0908,
        "open_interest": 4065,
        "volume": 229,
        "display_pl": 117.00,
        "buyback_trigger": "Only buy back early if you expect a major upside break before expiration or the mark falls below $0.25.",
        "roll_trigger": "Roll if ONDS closes above $15 or call delta rises above 0.45.",
        "action_now": "Hold. Upside is capped only above $20 on 200 shares, not near current price.",
    },
]

TRADE_LADDER = [
    ["NVDA", "Now", "121.67", "Trim", 5, "Reduce single-name concentration without exiting the winner."],
    ["NVDA", "Strength", "130.00", "Trim", 5, "Bring NVDA closer to 40% target if momentum continues."],
    ["NVDA", "Stretch", "140.00", "Trim", 5, "Only if still overweight after prior trims."],
    ["ONDS", "Pre/Post earnings spike", "11.50", "Trim uncovered shares", 53.89, "Sell only the uncovered block first; preserve 300 shares for short calls."],
    ["ONDS", "Strength", "13.50", "Trim uncovered shares", 50, "Take risk off into volatility rather than waiting for a full reversal."],
    ["ONDS", "Strength", "15.00", "Trim uncovered shares", 50, "Finish Phase 1 de-risking if price momentum remains constructive."],
    ["ONDS 06/18/26 $20C", "Decay", "0.12 premium", "Buy back", 1, "Close if almost all premium is gone and you want to rewrite at a better strike."],
    ["ONDS 09/18/26 $20C", "Breakout", "0.45 delta", "Roll", 2, "Protect upside if ONDS starts trending materially higher."],
    ["VOO", "Buy on dip", "545.00", "Buy", 1, "Deploy trim proceeds into the core index position on weakness."],
    ["VOO", "Buy on deeper dip", "530.00", "Buy", 1, "Scale in rather than chase strength."],
    ["SLS", "Simplify", "2.25", "Exit", 57.78, "Use liquidity pops to fully exit the small biotech flyer."],
]

SOURCES = [
    ("Current prices", "https://finance.yahoo.com", "Prices used in workbook: VOO 552.75, ONDS 10.16, NVDA 121.67, SLS 1.75."),
    ("ONDS official release", "https://ir.ondas.com/news-events/press-releases/detail/312/ondas-holdings-reports-fourth-quarter-and-full-year-2025-financial-results", "Ondas reported Q4 2025 and full-year 2025 results on March 12, 2026."),
    ("SLS official release", "https://www.sellaslifesciences.com/investors/news/News-Details/2026/SELLAS-Life-Sciences-Reports-Fourth-Quarter-and-Full-Year-2025-Financial-Results-and-Provides-Corporate-Update/default.aspx", "SELLAS reported Q4/full-year 2025 results and corporate update on March 12, 2026."),
    ("ONDS options chain", "https://www.barchart.com/stocks/quotes/ONDS/options?expiration=2026-06-18", "June 18, 2026 $20 call midpoint approximately $0.52 on March 13, 2026."),
    ("ONDS options chain", "https://www.barchart.com/stocks/quotes/ONDS/options?expiration=2026-09-18", "September 18, 2026 $20 call midpoint approximately $1.20 on March 13, 2026."),
]


def apply_sheet_theme(ws) -> None:
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="203864")
    for cell in row:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def currency(cell) -> None:
    cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'


def pct(cell) -> None:
    cell.number_format = '0.0%'


def whole(cell) -> None:
    cell.number_format = '#,##0.00;[Red](#,##0.00);-'


def add_title(ws, title: str, subtitle: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")


def build_summary(ws) -> None:
    add_title(ws, "Portfolio Management Plan", f"{ACCOUNT_NAME} | {AS_OF_LABEL}")
    ws["A4"] = "Goal"
    ws["B4"] = PORTFOLIO_GOAL
    ws["A5"] = "Recommended mix while ONDS calls remain open"
    ws["B5"] = "VOO 30% | NVDA 38% | ONDS 22% minimum | SLS 0% | Cash 10%"
    ws["A6"] = "Recommended mix after ONDS calls are closed/expire"
    ws["B6"] = "VOO 32% | NVDA 40% | ONDS 18% | SLS 0% | Cash 10%"
    ws["A8"] = "Highest-priority actions"
    ws["A9"] = "1. Use NVDA trims to fund VOO and cash reserve."
    ws["A10"] = "2. Do not cut ONDS below 300 shares while 3 covered calls remain open."
    ws["A11"] = "3. Hold the ONDS $20 calls for now; they are still comfortably out of the money."
    ws["A12"] = "4. Exit SLS on strength or simplify now because it is too small to move the portfolio."
    ws["A14"] = "Not financial advice"
    ws["B14"] = "This workbook is a disciplined trade-planning tool, not a guarantee of returns."
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 90


def build_holdings(ws) -> None:
    add_title(ws, "Holdings Snapshot", AS_OF_LABEL)
    headers = [
        "Ticker",
        "Name",
        "Shares",
        "Last Price",
        "Market Value",
        "Current Weight",
        "App P/L",
        "Covered Min Shares",
        "Phase 1 Target Weight",
        "Phase 1 Target Shares",
        "Phase 1 Share Delta",
        "Phase 2 Target Weight",
        "Phase 2 Target Shares",
        "Phase 2 Share Delta",
        "Note",
    ]
    start_row = 4
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(start_row, col_idx, value)
    style_header(ws[start_row])

    total_value_row = start_row + len(HOLDINGS) + 1
    for idx, item in enumerate(HOLDINGS, start=start_row + 1):
        ws.cell(idx, 1, item["ticker"])
        ws.cell(idx, 2, item["name"])
        ws.cell(idx, 3, item["shares"])
        ws.cell(idx, 4, item["price"])
        ws.cell(idx, 5, f"=C{idx}*D{idx}")
        ws.cell(idx, 6, f"=E{idx}/$E${total_value_row}")
        ws.cell(idx, 7, item["display_pl"])
        ws.cell(idx, 8, item["covered_min_shares"])
        ws.cell(idx, 9, item["target_phase1"])
        if item["covered_min_shares"]:
            ws.cell(idx, 10, f"=MAX(H{idx},($E${total_value_row}*I{idx})/D{idx})")
        else:
            ws.cell(idx, 10, f"=($E${total_value_row}*I{idx})/D{idx}")
        ws.cell(idx, 11, f"=J{idx}-C{idx}")
        ws.cell(idx, 12, item["target_phase2"])
        ws.cell(idx, 13, f"=($E${total_value_row}*L{idx})/D{idx}")
        ws.cell(idx, 14, f"=M{idx}-C{idx}")
        ws.cell(idx, 15, item["action_now"])

    ws.cell(total_value_row, 4, "Total")
    ws.cell(total_value_row, 5, f"=SUM(E{start_row + 1}:E{total_value_row - 1})")
    ws.cell(total_value_row + 1, 4, "Target cash")
    ws.cell(total_value_row + 1, 5, f"=E{total_value_row}*10%")

    for row in range(start_row + 1, total_value_row + 2):
        whole(ws.cell(row, 3))
        currency(ws.cell(row, 4))
        currency(ws.cell(row, 5))
        pct(ws.cell(row, 6))
        currency(ws.cell(row, 7))
        whole(ws.cell(row, 8))
        pct(ws.cell(row, 9))
        whole(ws.cell(row, 10))
        whole(ws.cell(row, 11))
        pct(ws.cell(row, 12))
        whole(ws.cell(row, 13))
        whole(ws.cell(row, 14))

    widths = [10, 28, 10, 12, 14, 14, 12, 16, 18, 20, 18, 18, 20, 18, 64]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A5"


def build_options(ws) -> None:
    add_title(ws, "ONDS Covered Call Management", AS_OF_LABEL)
    headers = [
        "Description",
        "Expiration",
        "Contracts",
        "Strike",
        "Bid",
        "Ask",
        "Mid",
        "Last",
        "Delta",
        "IV",
        "Open Interest",
        "Volume",
        "App P/L",
        "Buyback Cost",
        "Action Now",
        "Buyback Trigger",
        "Roll Trigger",
    ]
    start_row = 4
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(start_row, col_idx, value)
    style_header(ws[start_row])

    underlying_price = next(item["price"] for item in HOLDINGS if item["ticker"] == "ONDS")
    ws["A2"] = "ONDS spot"
    ws["B2"] = underlying_price
    currency(ws["B2"])

    for idx, item in enumerate(OPTIONS, start=start_row + 1):
        ws.cell(idx, 1, item["description"])
        ws.cell(idx, 2, item["expiration"])
        ws.cell(idx, 3, item["contracts"])
        ws.cell(idx, 4, item["strike"])
        ws.cell(idx, 5, item["bid"])
        ws.cell(idx, 6, item["ask"])
        ws.cell(idx, 7, item["mid"])
        ws.cell(idx, 8, item["last"])
        ws.cell(idx, 9, item["delta"])
        ws.cell(idx, 10, item["iv"])
        ws.cell(idx, 11, item["open_interest"])
        ws.cell(idx, 12, item["volume"])
        ws.cell(idx, 13, item["display_pl"])
        ws.cell(idx, 14, f"=ABS(C{idx})*G{idx}*100")
        ws.cell(idx, 15, item["action_now"])
        ws.cell(idx, 16, item["buyback_trigger"])
        ws.cell(idx, 17, item["roll_trigger"])

    for row in range(start_row + 1, start_row + 1 + len(OPTIONS)):
        for col in [4, 5, 6, 7, 8, 13, 14]:
            currency(ws.cell(row, col))
        pct(ws.cell(row, 9))
        pct(ws.cell(row, 10))

    ws["A8"] = "Combined estimated buyback cost now"
    ws["B8"] = "=SUM(N5:N6)"
    currency(ws["B8"])
    ws["A9"] = "Why this matters"
    ws["B9"] = "Closing all 3 calls would cost roughly $292 at mid prices, so early buybacks should be catalyst-driven, not automatic."

    widths = [20, 14, 10, 10, 10, 10, 10, 10, 10, 10, 14, 10, 10, 14, 42, 54, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A5"


def build_trade_ladder(ws) -> None:
    add_title(ws, "Trim And Buy Ladder", "Use limit orders and avoid chasing.")
    headers = ["Asset", "Setup", "Trigger", "Action", "Quantity", "Reason"]
    start_row = 4
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(start_row, col_idx, value)
    style_header(ws[start_row])
    for idx, row in enumerate(TRADE_LADDER, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value)
    ws.freeze_panes = "A5"
    widths = [22, 20, 16, 18, 12, 72]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_sources(ws) -> None:
    add_title(ws, "Sources", "URLs included so you can refresh the workbook later.")
    headers = ["Type", "URL", "Note"]
    start_row = 4
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(start_row, col_idx, value)
    style_header(ws[start_row])
    for idx, row in enumerate(SOURCES, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(idx, col_idx, value)
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96
    ws.column_dimensions["C"].width = 72


def add_comments(ws) -> None:
    ws["B5"].comment = Comment(
        "Phase 1 keeps ONDS at or above 300 shares because the account has 3 short covered calls outstanding.",
        "Codex",
    )
    ws["B6"].comment = Comment(
        "Phase 2 applies after the ONDS covered calls are closed, rolled, or expire.",
        "Codex",
    )


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    build_summary(summary)
    add_comments(summary)

    holdings = wb.create_sheet("Holdings")
    build_holdings(holdings)

    options = wb.create_sheet("Options")
    build_options(options)

    ladder = wb.create_sheet("Trade Ladder")
    build_trade_ladder(ladder)

    sources = wb.create_sheet("Sources")
    build_sources(sources)

    for ws in wb.worksheets:
        apply_sheet_theme(ws)

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
